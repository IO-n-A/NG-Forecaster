#!/usr/bin/env python3
"""Generate structured markdown companions for files under data/new.

Outputs:
- PDF transcription: <file>.md
- CSV summary: <file>.md
- XLSX summary: <file>.md

For CSV/XLSX summaries, include whitepaper-aligned implementation factor lineout.
"""

from __future__ import annotations

import argparse
from collections import Counter
import datetime as dt
import math
import re
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader


MAX_CELL_LEN = 140
MAX_SAMPLE_ROWS = 3
MAX_NUMERIC_COLS = 12
MAX_COLUMNS_PROFILED = 60
MAX_OBJECT_VALUES = 80
MAX_SHEET_PREVIEW_ROWS = 2
MAX_SHEET_PREVIEW_COLS = 8
MAX_INFERRED_COLUMNS = 8


PDF_NOISE_PATTERNS = [
    r"^\s*page\s*[|:]?\s*\d+\s*$",
    r"^\s*\d+\s*\|\s*page\s*\|?\s*\d*\s*$",
    r"^\s*u\.s\. energy information administration\s*\|.*$",
    r"^\s*iea\.?\s*cc by.*$",
    r"^\s*international energy agency\s*$",
    r"^\s*website:\s*www\.[^\s]+$",
    r"^\s*www\.[^\s]+$",
    r"^\s*the iea examines the full spectrum of energy issues.*$",
    r"^\s*this publication and any map included herein.*$",
    r"^\s*revised version,?\s*[a-z]+\s*\d{4}.*$",
    r"^\s*information notice found at:?\s*$",
]

COUNTRY_LIST_ENTRIES = {
    "Argentina",
    "Australia",
    "Austria",
    "Belgium",
    "Brazil",
    "Canada",
    "China",
    "Czech Republic",
    "Denmark",
    "Egypt",
    "Estonia",
    "Finland",
    "France",
    "Germany",
    "Greece",
    "Hungary",
    "India",
    "Indonesia",
    "Ireland",
    "Italy",
    "Japan",
    "Kenya",
    "Korea",
    "Latvia",
    "Lithuania",
    "Luxembourg",
    "Mexico",
    "Morocco",
    "Netherlands",
    "New Zealand",
    "Norway",
    "Poland",
    "Portugal",
    "Republic of Türkiye",
    "Senegal",
    "Singapore",
    "Slovak Republic",
    "South Africa",
    "Spain",
    "Sweden",
    "Switzerland",
    "Thailand",
    "Ukraine",
    "United Kingdom",
    "United States",
}


FACTOR_CATALOG: list[dict[str, object]] = [
    {
        "name": "Production level/state",
        "keywords": ["production", "dry", "marketed", "ngpr", "supply", "output", "shale"],
        "features": ["prod_level", "prod_mom", "prod_yoy", "prod_3m_avg"],
        "whitepaper_link": "N2 lag-safe feature family for monthly production state.",
    },
    {
        "name": "Storage level/state",
        "keywords": ["storage", "working gas", "stock", "inventory", "stor"],
        "features": ["stor_last", "stor_mean_4w", "stor_slope_4w", "stor_mom_change"],
        "whitepaper_link": "N2 baseline storage drivers (explicitly listed in whitepaper v3).",
    },
    {
        "name": "Price level/state",
        "keywords": ["henry", "hub", "price", "spot", "futures", "hh"],
        "features": ["hh_mtd_mean", "hh_last", "hh_vol_7d", "hh_diff_1d"],
        "whitepaper_link": "N2 baseline price factors (explicitly listed in whitepaper v3).",
    },
    {
        "name": "Drilling/productivity",
        "keywords": [
            "rig",
            "wells",
            "drilling",
            "duc",
            "completed",
            "legacy",
            "productivity",
        ],
        "features": ["rig_count_lag1", "duc_delta", "new_well_output", "legacy_decline_rate"],
        "whitepaper_link": "N2 feature expansion; supports N4 challenger robustness on regime shifts.",
    },
    {
        "name": "Demand/consumption",
        "keywords": ["demand", "consumption", "residential", "commercial", "industrial", "power"],
        "features": ["sector_demand_share", "power_burn_proxy", "demand_mom", "demand_yoy"],
        "whitepaper_link": "Tier-2 optional exogenous family; keep N1 as-of and lag checks enforced.",
    },
    {
        "name": "LNG/exports/flows",
        "keywords": ["lng", "export", "import", "feedgas", "pipeline", "net export"],
        "features": ["lng_feedgas_level", "net_exports", "pipeline_flow_delta"],
        "whitepaper_link": "Useful exogenous regime-state factors for N4 fusion/challenger diagnostics.",
    },
    {
        "name": "Weather/seasonality",
        "keywords": ["weather", "temperature", "hdd", "cdd", "degree day", "heating", "cooling"],
        "features": ["hdd_anom", "cdd_anom", "temp_zscore"],
        "whitepaper_link": "Supports deterministic monthly aggregation while preserving N1 admissibility.",
    },
    {
        "name": "Macro/economic",
        "keywords": ["gdp", "industrial production", "economy", "manufacturing", "income"],
        "features": ["ip_index_proxy", "macro_growth_proxy"],
        "whitepaper_link": "Optional Tier-2 enrichments; run through N6 ablation gate before adoption.",
    },
]


@dataclass
class RunStats:
    pdf_total: int = 0
    csv_total: int = 0
    xlsx_total: int = 0
    ok: int = 0
    failed: int = 0


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


def rel(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def file_size(path: Path) -> int:
    return path.stat().st_size if path.exists() else 0


def clean_pdf_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # Light de-hyphenation only when the next line starts with lowercase letters.
    text = re.sub(r"([A-Za-z])-\n([a-z])", r"\1\2", text)
    lines = [ln.rstrip() for ln in text.split("\n")]

    cleaned_lines: list[str] = []
    blank_streak = 0
    for line in lines:
        if line.strip():
            blank_streak = 0
            cleaned_lines.append(line)
        else:
            blank_streak += 1
            if blank_streak <= 1:
                cleaned_lines.append("")
    return "\n".join(cleaned_lines).strip()


def is_pdf_noise_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    low = stripped.lower()
    if low.startswith("the iea examines the full spectrum of energy issues"):
        return True
    if low.startswith("this publication and any map included herein"):
        return True
    if low.startswith("revised version"):
        return True
    if "information notice found at" in low:
        return True
    for pattern in PDF_NOISE_PATTERNS:
        if re.match(pattern, low):
            return True
    if stripped in {"Source: IEA.", "Source: IEA"}:
        return True
    return False


def remove_member_country_blocks(lines: list[str]) -> list[str]:
    out: list[str] = []
    i = 0
    while i < len(lines):
        stripped = lines[i].strip()
        is_block_start = bool(
            re.match(r"^(?:iea\s+)?member countries:?\s*$", stripped, flags=re.IGNORECASE)
            or re.match(r"^(?:iea\s+)?association countries:?\s*$", stripped, flags=re.IGNORECASE)
        )
        if not is_block_start:
            out.append(lines[i])
            i += 1
            continue

        i += 1
        while i < len(lines):
            candidate = lines[i].strip()
            if not candidate:
                i += 1
                continue
            if candidate in COUNTRY_LIST_ENTRIES:
                i += 1
                continue
            candidate_low = candidate.lower()
            if candidate_low.startswith("the european commission"):
                i += 1
                continue
            if "participates in the work of the iea" in candidate_low:
                i += 1
                continue
            if candidate_low == "international energy agency":
                i += 1
                continue
            # Keep skipping short title-case list items while inside this block.
            if len(candidate) <= 32 and re.fullmatch(r"[A-Za-z][A-Za-z.\-']+( [A-Za-z][A-Za-z.\-']+){0,4}", candidate):
                i += 1
                continue
            break

        if out and out[-1].strip():
            out.append("")

    return out


def normalize_heading(line: str) -> str:
    if line.isupper():
        words = []
        for token in line.lower().split():
            if token in {"and", "or", "of", "the", "to", "for", "in", "on"}:
                words.append(token)
            else:
                words.append(token.capitalize())
        return " ".join(words)
    return line


def is_heading_candidate(line: str) -> bool:
    if len(line) < 4 or len(line) > 90:
        return False
    if line.endswith("."):
        return False
    alpha = [ch for ch in line if ch.isalpha()]
    if len(alpha) < 4:
        return False
    words = [w for w in line.split() if any(c.isalpha() for c in w)]
    if not words or len(words) > 14:
        return False
    upper_ratio = sum(1 for c in alpha if c.isupper()) / len(alpha)
    title_ratio = sum(1 for w in words if w[0].isupper()) / len(words)
    if upper_ratio >= 0.75:
        return True
    if title_ratio >= 0.85 and len(line) <= 75:
        return True
    if re.match(r"^(chapter|section|table|figure)\b", line.strip(), flags=re.IGNORECASE):
        return True
    return False


def is_list_item_candidate(line: str) -> bool:
    return bool(
        re.match(r"^\s*[-*•]\s+", line)
        or re.match(r"^\s*\(?\d{1,2}[.)]\s+", line)
    )


def collapse_blank_lines(lines: list[str]) -> list[str]:
    out: list[str] = []
    blank_streak = 0
    for line in lines:
        if line.strip():
            blank_streak = 0
            out.append(line)
        else:
            blank_streak += 1
            if blank_streak <= 1:
                out.append("")
    while out and not out[-1].strip():
        out.pop()
    return out


def sanitize_pdf_lines(text: str) -> list[str]:
    cleaned = clean_pdf_text(text)
    if not cleaned:
        return []

    lines = [ln.strip() for ln in cleaned.splitlines()]
    lines = remove_member_country_blocks(lines)
    lines = [ln for ln in lines if not is_pdf_noise_line(ln)]
    lines = collapse_blank_lines(lines)
    return lines


def recurring_page_header_lines(page_lines: list[list[str]]) -> set[str]:
    if not page_lines:
        return set()

    counts: Counter[str] = Counter()
    for lines in page_lines:
        seen: set[str] = set()
        for line in lines[:5]:
            normalized = normalize_text_for_matching(line)
            if len(normalized) < 8:
                continue
            if normalized in seen:
                continue
            seen.add(normalized)
        for normalized in seen:
            counts[normalized] += 1

    threshold = max(4, len(page_lines) // 10)
    return {
        text
        for text, cnt in counts.items()
        if cnt >= threshold
    }


def render_pdf_page_markdown(lines: list[str]) -> str:
    if not lines:
        return "(No extractable text on this page.)"

    out_blocks: list[str] = []
    para: list[str] = []

    def append_block(block: str) -> None:
        text = block.strip()
        if not text:
            return
        if out_blocks and normalize_text_for_matching(out_blocks[-1]) == normalize_text_for_matching(text):
            return
        out_blocks.append(text)

    def flush_para() -> None:
        if not para:
            return
        append_block(" ".join(para))
        para.clear()

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            flush_para()
            continue
        if is_heading_candidate(line):
            flush_para()
            append_block(f"#### {normalize_heading(line)}")
            continue
        if is_list_item_candidate(line):
            flush_para()
            cleaned_item = re.sub(r"^\s*[-*•]\s+", "", line).strip()
            append_block(f"- {cleaned_item}")
            continue
        para.append(line)

    flush_para()
    if not out_blocks:
        return "(No extractable text on this page.)"
    return "\n\n".join(out_blocks)


def short_text(value: object, max_len: int = MAX_CELL_LEN) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) > max_len:
        return text[: max_len - 3] + "..."
    return text


def md_escape(text: str) -> str:
    return text.replace("|", r"\|")


def md_table(headers: list[str], rows: list[list[object]]) -> str:
    if not rows:
        return "_No rows to display._"
    head = "| " + " | ".join(md_escape(h) for h in headers) + " |"
    sep = "| " + " | ".join("---" for _ in headers) + " |"
    body_lines = []
    for row in rows:
        values = [md_escape(short_text(v)) for v in row]
        if len(values) < len(headers):
            values.extend([""] * (len(headers) - len(values)))
        body_lines.append("| " + " | ".join(values[: len(headers)]) + " |")
    return "\n".join([head, sep, *body_lines])


def normalize_text_for_matching(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def keyword_hits(corpus: str, keywords: Iterable[str]) -> list[str]:
    hits: list[str] = []
    normalized = normalize_text_for_matching(corpus)
    for kw in keywords:
        needle = normalize_text_for_matching(kw)
        if not needle:
            continue
        if " " in needle:
            if needle in normalized:
                hits.append(kw)
        elif re.search(rf"\b{re.escape(needle)}\b", normalized):
            hits.append(kw)
    return hits


def build_factor_lineout(corpus: str) -> list[list[str]]:
    rows: list[list[str]] = []
    for factor in FACTOR_CATALOG:
        hits = keyword_hits(corpus, factor["keywords"])
        if not hits:
            continue
        rows.append(
            [
                str(factor["name"]),
                ", ".join(sorted(set(hits)))[:180],
                ", ".join(factor["features"]),  # type: ignore[arg-type]
                str(factor["whitepaper_link"]),
            ]
        )
    return rows


def parse_dates_from_series(s: pd.Series) -> tuple[int, str, str]:
    parsed = pd.to_datetime(s, errors="coerce")
    valid = parsed.dropna()
    if valid.empty:
        return 0, "", ""
    return int(valid.shape[0]), str(valid.min().date()), str(valid.max().date())


def looks_like_compact_date(value: object) -> bool:
    text = short_text(value, max_len=16)
    text = text.replace(".0", "").strip()
    return bool(re.fullmatch(r"\d{6}|\d{8}", text))


def to_compact_date_string(value: object) -> str:
    return short_text(value, max_len=16).replace(".0", "").strip()


def write_pdf_markdown(pdf_path: Path, root: Path) -> None:
    out_path = pdf_path.with_suffix(".md")
    reader = PdfReader(str(pdf_path))
    page_count = len(reader.pages)
    page_lines: list[list[str]] = []
    for page in reader.pages:
        raw_text = page.extract_text() or ""
        page_lines.append(sanitize_pdf_lines(raw_text))
    recurring_headers = recurring_page_header_lines(page_lines)

    with out_path.open("w", encoding="utf-8") as f:
        f.write(f"# PDF Transcription: {pdf_path.name}\n\n")
        f.write("## Source Metadata\n")
        f.write(f"- file: `{rel(pdf_path, root)}`\n")
        f.write(f"- size_bytes: `{file_size(pdf_path)}`\n")
        f.write(f"- page_count: `{page_count}`\n")
        f.write("- extraction_engine: `pypdf.PdfReader.extract_text`\n")
        f.write("- extraction_note: `OCR was not applied; scanned-image pages may have sparse text.`\n")
        f.write(f"- generated_at_utc: `{utc_now()}`\n\n")
        f.write("## Full Transcription\n\n")

        for idx, lines in enumerate(page_lines, start=1):
            filtered_lines: list[str] = []
            for line in lines:
                normalized = normalize_text_for_matching(line)
                if normalized and normalized in recurring_headers:
                    continue
                filtered_lines.append(line)
            filtered_lines = collapse_blank_lines(filtered_lines)
            f.write(f"### Page {idx}\n")
            f.write("\n")
            f.write(render_pdf_page_markdown(filtered_lines))
            f.write("\n\n")


def read_csv_with_fallback(path: Path) -> tuple[pd.DataFrame, str]:
    encodings = ["utf-8", "utf-8-sig", "latin-1"]
    last_error: Exception | None = None
    for enc in encodings:
        try:
            df = pd.read_csv(path, encoding=enc, low_memory=False)
            return df, enc
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            continue
    raise RuntimeError(f"failed to read CSV with fallback encodings: {last_error}")


def dataframe_column_profile(df: pd.DataFrame) -> list[list[object]]:
    rows: list[list[object]] = []
    for col in list(df.columns)[:MAX_COLUMNS_PROFILED]:
        s = df[col]
        missing = int(s.isna().sum())
        missing_pct = (missing / max(1, len(df))) * 100.0
        distinct = int(s.nunique(dropna=True))
        rows.append([str(col), str(s.dtype), len(s) - missing, missing, f"{missing_pct:.2f}%", distinct])
    return rows


def dataframe_date_profile(df: pd.DataFrame) -> list[list[str]]:
    rows: list[list[str]] = []
    for col in list(df.columns)[:MAX_COLUMNS_PROFILED]:
        s = df[col].dropna()
        if s.empty:
            continue
        probe = s.head(2000)
        if pd.api.types.is_datetime64_any_dtype(probe):
            parsed_count, min_dt, max_dt = parse_dates_from_series(probe)
        elif pd.api.types.is_object_dtype(probe) or pd.api.types.is_string_dtype(probe):
            parsed_count, min_dt, max_dt = parse_dates_from_series(probe)
        else:
            values = probe.tolist()
            compact_ratio = sum(1 for v in values if looks_like_compact_date(v)) / max(1, len(values))
            if compact_ratio < 0.8:
                continue
            compact_series = pd.Series([to_compact_date_string(v) for v in values])
            parsed_count, min_dt, max_dt = parse_dates_from_series(compact_series)
        if parsed_count < max(5, int(min(len(s), 2000) * 0.5)):
            continue
        rows.append([str(col), str(parsed_count), min_dt, max_dt])
    return rows[:20]


def dataframe_numeric_profile(df: pd.DataFrame) -> list[list[str]]:
    rows: list[list[str]] = []
    for col in list(df.columns)[:MAX_COLUMNS_PROFILED]:
        s = pd.to_numeric(df[col], errors="coerce")
        valid = s.dropna()
        if valid.shape[0] < 5:
            continue
        rows.append(
            [
                str(col),
                str(valid.shape[0]),
                f"{float(valid.mean()):.6g}",
                f"{float(valid.std(ddof=0)):.6g}",
                f"{float(valid.min()):.6g}",
                f"{float(valid.max()):.6g}",
            ]
        )
        if len(rows) >= MAX_NUMERIC_COLS:
            break
    return rows


def dataframe_sample_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "_No data rows._"
    sample = df.head(MAX_SAMPLE_ROWS).copy()
    columns = [str(c) for c in sample.columns]
    rows: list[list[object]] = []
    for _, row in sample.iterrows():
        rows.append([short_text(row[c]) for c in sample.columns])
    return md_table(columns, rows)


def dataframe_corpus(df: pd.DataFrame, name_hint: str) -> str:
    parts: list[str] = [name_hint]
    parts.extend(str(c) for c in df.columns)
    for col in df.select_dtypes(include=["object", "string"]).columns[:20]:
        vals = (
            df[col]
            .dropna()
            .astype(str)
            .head(MAX_OBJECT_VALUES)
            .map(lambda x: short_text(x, max_len=60))
            .tolist()
        )
        for value in vals:
            if not re.search(r"[A-Za-z]", value):
                continue
            if re.fullmatch(r"[0-9a-f]{16,}", value.lower()):
                continue
            parts.append(value)
    return " ".join(parts).lower()


def write_csv_markdown(csv_path: Path, root: Path) -> None:
    out_path = csv_path.with_suffix(".md")
    df, encoding = read_csv_with_fallback(csv_path)

    factor_rows = build_factor_lineout(dataframe_corpus(df, csv_path.stem))
    column_profile = dataframe_column_profile(df)
    date_profile = dataframe_date_profile(df)
    numeric_profile = dataframe_numeric_profile(df)

    with out_path.open("w", encoding="utf-8") as f:
        f.write(f"# CSV Summary: {csv_path.name}\n\n")
        f.write("## Source Metadata\n")
        f.write(f"- file: `{rel(csv_path, root)}`\n")
        f.write(f"- size_bytes: `{file_size(csv_path)}`\n")
        f.write(f"- encoding_used: `{encoding}`\n")
        f.write(f"- rows: `{df.shape[0]}`\n")
        f.write(f"- columns: `{df.shape[1]}`\n")
        f.write(f"- generated_at_utc: `{utc_now()}`\n\n")

        f.write("## Column Profile\n")
        f.write(
            md_table(
                ["column", "dtype", "non_null", "missing", "missing_pct", "distinct_values"],
                column_profile,
            )
        )
        f.write("\n\n")

        f.write("## Date-like Columns\n")
        f.write(md_table(["column", "parsed_rows", "min_date", "max_date"], date_profile))
        f.write("\n\n")

        f.write("## Numeric Column Stats (sampled)\n")
        f.write(md_table(["column", "valid_n", "mean", "std", "min", "max"], numeric_profile))
        f.write("\n\n")

        f.write("## Sample Rows\n")
        f.write(dataframe_sample_table(df))
        f.write("\n\n")

        f.write("## Whitepaper v3 Implementation Lineout\n")
        f.write(
            "Candidate values/factors for `docs/method/whitepaper_v3.md` based on "
            "observed columns/content hints.\n\n"
        )
        if factor_rows:
            f.write(
                md_table(
                    [
                        "factor_family",
                        "evidence_keywords_found",
                        "candidate_features",
                        "implementation_link",
                    ],
                    factor_rows,
                )
            )
        else:
            f.write(
                "No strong domain keyword hits were detected. Treat this file as metadata/supporting "
                "lineage input for N1/N3 reproducibility artifacts.\n"
            )
        f.write("\n\n")

        f.write("## Guardrail Notes\n")
        f.write("- Enforce N1 as-of cutoff and target lag policy before feature materialization.\n")
        f.write("- Run N3 deterministic preprocessing prior to any model ingestion.\n")
        f.write("- Validate factor additions via N6 ablation stages (`B0`..`B4`) before adoption.\n")


@dataclass
class SheetSummary:
    name: str
    max_row: int
    max_col: int
    header_row_guess: int
    inferred_columns: list[str]
    sample_rows: list[list[str]]
    date_like_count: int
    date_min: str
    date_max: str
    numeric_count: int


def maybe_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if re.fullmatch(r"-?\d+(\.\d+)?", text):
        try:
            return float(text)
        except ValueError:
            return None
    return None


def sheet_header_guess(sample_rows: list[list[str]]) -> int:
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(sample_rows, start=1):
        score = sum(1 for cell in row if cell and not re.fullmatch(r"-?\d+(\.\d+)?", cell))
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx


def summarize_workbook(path: Path) -> tuple[list[SheetSummary], str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    summaries: list[SheetSummary] = []
    corpus_parts: list[str] = [path.stem]

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = int(ws.max_row or 0)
            max_col = int(ws.max_column or 0)

            sample_rows: list[list[str]] = []
            date_values: list[dt.date] = []
            numeric_count = 0

            for row in ws.iter_rows(
                min_row=1,
                max_row=min(max_row, 80),
                min_col=1,
                max_col=min(max_col, MAX_SHEET_PREVIEW_COLS),
                values_only=True,
            ):
                txt_row = [short_text(v, max_len=80) for v in row]
                if any(cell for cell in txt_row):
                    sample_rows.append(txt_row)
                for value in row:
                    if isinstance(value, dt.datetime):
                        date_values.append(value.date())
                    elif isinstance(value, dt.date):
                        date_values.append(value)
                    else:
                        as_float = maybe_float(value)
                        if as_float is not None:
                            numeric_count += 1
                if len(sample_rows) >= 10 and len(date_values) >= 80 and numeric_count >= 220:
                    break

            header_guess = sheet_header_guess(sample_rows[:10]) if sample_rows else 0
            inferred_columns: list[str] = []
            if sample_rows and 1 <= header_guess <= len(sample_rows):
                header_row = sample_rows[header_guess - 1]
                inferred_columns = [short_text(c, max_len=36) for c in header_row if c][:MAX_INFERRED_COLUMNS]

            for row in sample_rows[:8]:
                corpus_parts.extend(cell for cell in row if cell)
            corpus_parts.append(sheet_name)

            if date_values:
                date_min = min(date_values).isoformat()
                date_max = max(date_values).isoformat()
            else:
                date_min = ""
                date_max = ""

            summaries.append(
                SheetSummary(
                    name=sheet_name,
                    max_row=max_row,
                    max_col=max_col,
                    header_row_guess=header_guess,
                    inferred_columns=inferred_columns,
                    sample_rows=sample_rows[:MAX_SHEET_PREVIEW_ROWS],
                    date_like_count=len(date_values),
                    date_min=date_min,
                    date_max=date_max,
                    numeric_count=numeric_count,
                )
            )
    finally:
        wb.close()

    corpus = " ".join(corpus_parts).lower()
    return summaries, corpus


def sheet_row_preview(row: list[str]) -> str:
    cells = [short_text(c, max_len=48) for c in row if c]
    if not cells:
        return "(empty)"
    return " | ".join(cells[:MAX_SHEET_PREVIEW_COLS])


def write_xlsx_markdown(xlsx_path: Path, root: Path) -> None:
    out_path = xlsx_path.with_suffix(".md")
    summaries, corpus = summarize_workbook(xlsx_path)
    factor_rows = build_factor_lineout(corpus)

    with out_path.open("w", encoding="utf-8") as f:
        f.write(f"# XLSX Summary: {xlsx_path.name}\n\n")
        f.write("## Source Metadata\n")
        f.write(f"- file: `{rel(xlsx_path, root)}`\n")
        f.write(f"- size_bytes: `{file_size(xlsx_path)}`\n")
        f.write(f"- sheet_count: `{len(summaries)}`\n")
        f.write("- extraction_engine: `openpyxl(read_only=True, data_only=True)`\n")
        f.write(f"- generated_at_utc: `{utc_now()}`\n\n")

        overview_rows = [
            [
                s.name,
                s.max_row,
                s.max_col,
                s.header_row_guess if s.header_row_guess else "",
                s.date_like_count,
                s.date_min,
                s.date_max,
                s.numeric_count,
            ]
            for s in summaries
        ]
        f.write("## Workbook Overview\n")
        f.write(
            md_table(
                [
                    "sheet",
                    "max_row",
                    "max_col",
                    "header_row_guess",
                    "date_like_values_sampled",
                    "date_min",
                    "date_max",
                    "numeric_values_sampled",
                ],
                overview_rows,
            )
        )
        f.write("\n\n")

        f.write("## Sheet Highlights\n")
        if not summaries:
            f.write("_No sheets found._\n\n")
        for s in summaries:
            f.write(f"### Sheet: {s.name}\n")
            f.write(f"- dimensions: `{s.max_row} x {s.max_col}`\n")
            f.write(f"- inferred_columns: `{', '.join(s.inferred_columns) if s.inferred_columns else '(none inferred)'}`\n")
            if s.date_like_count:
                f.write(f"- sampled_date_range: `{s.date_min} -> {s.date_max}`\n")
            else:
                f.write("- sampled_date_range: `(none detected in sampled cells)`\n")
            f.write(f"- sampled_numeric_values: `{s.numeric_count}`\n")
            if s.sample_rows:
                f.write("- row_previews:\n")
                for i, row in enumerate(s.sample_rows, start=1):
                    f.write(f"  - r{i}: `{sheet_row_preview(row)}`\n")
            else:
                f.write("- row_previews: `(no non-empty sampled rows)`\n")
            f.write("\n")

        f.write("## Whitepaper v3 Implementation Lineout\n")
        f.write(
            "Candidate values/factors for `docs/method/whitepaper_v3.md` based on sheet names, "
            "header-like rows, and sampled content.\n\n"
        )
        if factor_rows:
            f.write(
                md_table(
                    [
                        "factor_family",
                        "evidence_keywords_found",
                        "candidate_features",
                        "implementation_link",
                    ],
                    factor_rows,
                )
            )
            f.write("\n\n")
        else:
            f.write(
                "No strong domain keyword hits were detected in sampled workbook content. "
                "Treat as supporting/lineage data unless downstream parser identifies series codes.\n\n"
            )

        f.write("## Guardrail Notes\n")
        f.write("- Enforce N1 as-of and lag eligibility in vintage builder before joining this source.\n")
        f.write("- Apply N3 preprocessing checks (missing/outlier policy) ahead of model stage.\n")
        f.write("- Promote any new factors only after N6 ablation evidence and DM/N5 policy checks.\n")


def discover_files(root: Path) -> tuple[list[Path], list[Path], list[Path]]:
    pdfs = sorted(p for p in root.rglob("*.pdf") if p.is_file())
    csvs = sorted(p for p in root.rglob("*.csv") if p.is_file())
    xlsxs = sorted(p for p in root.rglob("*.xlsx") if p.is_file())
    return pdfs, csvs, xlsxs


def process_all(root: Path, kinds: set[str]) -> RunStats:
    stats = RunStats()
    pdfs, csvs, xlsxs = discover_files(root)
    stats.pdf_total = len(pdfs)
    stats.csv_total = len(csvs)
    stats.xlsx_total = len(xlsxs)

    queue: list[tuple[str, Path]] = []
    if "pdf" in kinds:
        queue.extend(("pdf", p) for p in pdfs)
    if "csv" in kinds:
        queue.extend(("csv", p) for p in csvs)
    if "xlsx" in kinds:
        queue.extend(("xlsx", p) for p in xlsxs)

    total = len(queue)
    for idx, (kind, path) in enumerate(queue, start=1):
        print(f"[{idx}/{total}] processing {kind}: {rel(path, root)}")
        try:
            if kind == "pdf":
                write_pdf_markdown(path, root)
            elif kind == "csv":
                write_csv_markdown(path, root)
            else:
                write_xlsx_markdown(path, root)
            stats.ok += 1
        except Exception:  # noqa: BLE001
            stats.failed += 1
            err_path = path.with_suffix(path.suffix + ".error.log")
            err_path.write_text(traceback.format_exc(), encoding="utf-8")
            print(f"  failed -> {err_path.name}")

    return stats


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate structured markdown companions for PDFs/CSVs/XLSXs in data/new."
    )
    parser.add_argument("--root", type=Path, default=Path("data/new"), help="Root folder to scan")
    parser.add_argument(
        "--kinds",
        type=str,
        default="pdf,csv,xlsx",
        help="Comma-separated subset of kinds to process: pdf,csv,xlsx",
    )
    args = parser.parse_args()

    root = args.root.resolve()
    if not root.exists():
        raise SystemExit(f"Root path does not exist: {root}")

    kinds = {k.strip().lower() for k in args.kinds.split(",") if k.strip()}
    allowed = {"pdf", "csv", "xlsx"}
    if not kinds or not kinds.issubset(allowed):
        raise SystemExit(f"Invalid --kinds value: {args.kinds}; allowed: pdf,csv,xlsx")

    stats = process_all(root, kinds)
    print("\nSummary")
    print(f"- root: {root}")
    print(f"- pdf_total: {stats.pdf_total}")
    print(f"- csv_total: {stats.csv_total}")
    print(f"- xlsx_total: {stats.xlsx_total}")
    print(f"- succeeded: {stats.ok}")
    print(f"- failed: {stats.failed}")
    if stats.failed:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
